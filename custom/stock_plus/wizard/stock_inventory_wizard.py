from odoo import models, fields, _, api
from odoo.exceptions import UserError

import base64
import io
import openpyxl
import pickle


class StockInventoryWizard(models.TransientModel):
    _name = "stock.inventory.wizard"

    file = fields.Binary("Fichier Excel")
    filename = fields.Char("Nom de fichier")
    valid = fields.Boolean()
    message = fields.Html()
    data = fields.Binary()

    @api.onchange("file")
    def _onchange_file(self):
        self.valid = self.message = False

    def button_check(self):
        if not self.file or not self.filename:
            raise UserError("Veuillez charger un fichier Excel.")

        if not self.filename.lower().endswith(".xlsx"):
            raise UserError("Seuls les fichiers .xlsx sont pris en charge.")

        profile_ids = self.env["building.profile.assignment"].search([("user_id", "=", self.env.user.id), ("group_id.name", "=", "SOTASERV_CHEF_PROJET")])
        user_site_ids = profile_ids.mapped("site_id")

        file_content = base64.b64decode(self.file)
        file_stream = io.BytesIO(file_content)
        workbook = openpyxl.load_workbook(file_stream, data_only=True)
        sheet = workbook.active

        product_names = [row[0].value for row in sheet.iter_rows(min_row=2) if row[0].value]
        products = self.env["product.product"].search([("name", "in", product_names)])
        product_name_to_id = {product.name: product.id for product in products}

        headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        site_numbers = headers[1:]
        sites = self.env["building.site"].search([("number", "in", site_numbers)])

        invalid_products = [product for product in product_names if product not in product_name_to_id]
        invalid_sites = [site for site in site_numbers if site not in sites.mapped("number")]
        user_sites = sites & user_site_ids
        no_rights_sites = sites - user_site_ids

        site_number_to_id = {site.number: site.id for site in user_sites}

        # raise Exception(invalid_sites, user_sites, no_rights_sites)

        per_site = {site.id: {
            "site_number": site.number,
            "stock_location": site.warehouse_id.lot_stock_id.ids,
            "products": {},
        } for site in user_sites}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            product, *quantities = row
            if product not in product_name_to_id:
                continue

            for site, quantity in zip(site_numbers, quantities):
                if site not in site_number_to_id:
                    continue
                per_site[site_number_to_id[site]]["products"][product_name_to_id[product]] = quantity or 0

        message = """<div class="message">"""
        if bool(no_rights_sites):
            message += f"""
                <p class="red">Vous n'avez pas de droits d'accès à ces affaires</p>
                <ul>{"".join([f"<li>{site.number}</li>" for site in no_rights_sites])}</ul>
            """
        if bool(invalid_sites):
            message += f"""
                <p class="red">Affaires invalides</p>
                <ul>{"".join([f"<li>{site}</li>" for site in invalid_sites])}</ul>
            """
        if bool(invalid_products):
            message += f"""
                <p class="red">Articles invalides</p>
                <ul>{"".join([f"<li>{product}</li>" for product in invalid_products])}</ul>
            """
        if bool(per_site):
            self.valid = True
            self.data = base64.b64encode(pickle.dumps(per_site))
            result_length = len(per_site)
            p = f"""{result_length} {result_length > 1 and "inventaires seraient créés pour les affaires" or "inventaire serait créé pour l'affaire"}"""

            message += f"""
                <p class="green">{p}</p>
                <ul>{"".join([f"<li>{data['site_number']}</li>" for _, data in per_site.items()])}</ul>
            """
        message += "</div>"
        self.message = message

        return {
            "name": "Ajustement de l'inventaire – Excel",
            "type": "ir.actions.act_window",
            "view_mode": "form",
            "res_model": self._name,
            "res_id": self.id,
            "target": "new",
        }

    def button_create(self):
        per_site = pickle.loads(base64.b64decode(self.data))
        StockInventory = self.env["stock.inventory"]
        inventories = []

        for site, data in per_site.items():
            products = list(data["products"].keys())
            name = self.env["ir.sequence"].next_by_code("stock.inventory")
            inventories.append({
                "name": name,
                "site_id": site,
                "location_ids": data["stock_location"],
                "product_ids": products,
                "exhausted": True,
                "prefill_counted_quantity": "zero",
                "from_excel": True,
            })

        # raise Exception(inventories)

        inventories = StockInventory.create(inventories)

        for inventory in inventories:
            inventory.action_apply_inventory()
            site = inventory.site_id

            for line in inventory.line_ids:
                product = line.product_id   
                quantity = per_site[site.id]["products"][product.id]
                if line.theoretical_qty == quantity == 0: line.unlink()
                else: line.product_qty = quantity

            inventory.product_ids = inventory.line_ids.mapped("product_id")

        names = inventories.mapped("name")
        singleton = len(names) == 1
        notification_title = "Inventaire créé avec succès" if singleton else "Inventaires créés avec succès"
        notification_message = names[0] if singleton else ", ".join(names[:-1]) + " & " + names[-1]
        notification = {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "type": "success",
                "title": notification_title,
                "message": notification_message,
                "sticky": True,
            }
        }

        return {
            "type": "ir.actions.act_multi",
            "actions": [notification, {"type": "ir.actions.act_window_close"}],
        }